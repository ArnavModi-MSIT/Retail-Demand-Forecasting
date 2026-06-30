import pandas as pd
import numpy as np
import io


def generate_executive_summary(
    predictions_df: pd.DataFrame,
    inventory_df: pd.DataFrame,
    results_df: pd.DataFrame,
    model_name: str,
    horizon_days: int,
    future_forecast_df: pd.DataFrame = None,
) -> dict:
    """
    Build a reporting bundle with three metric groups:

    - evaluation_metrics: test-set model quality metrics
    - forecasting_metrics: future-horizon forecast outputs
    - business_metrics: operational inventory decisions

    The returned 'summary' dict remains a compatibility object for the existing
    executive summary UI and exports, but it only contains business-facing
    forecast and inventory metrics.
    """
    # ── Evaluation metrics (model quality on holdout/test data) ───────────────
    best = results_df[results_df["Model"] == model_name] if not results_df.empty else pd.DataFrame()
    mae = best["MAE"].values[0] if len(best) else None
    rmse = best["RMSE"].values[0] if len(best) else None
    mape = best["MAPE"].values[0] if len(best) else None

    naive_row = results_df[results_df["Model"].str.contains("Naive")] if not results_df.empty else pd.DataFrame()
    baseline_improvement = None
    if len(naive_row) and rmse:
        baseline_rmse = naive_row["RMSE"].values[0]
        if baseline_rmse:
            baseline_improvement = round(
                (baseline_rmse - rmse) / baseline_rmse * 100, 1
            )

    evaluation_metrics = {
        "model_name": model_name,
        "mae": mae,
        "rmse": rmse,
        "mape": mape,
        "baseline_improvement": baseline_improvement,
        "test_predictions_df": predictions_df.copy(),
        "results_df": results_df.copy(),
    }

    # ── Forecasting metrics (future-horizon demand outputs) ──────────────────
    forecast_source = future_forecast_df if future_forecast_df is not None and not future_forecast_df.empty else predictions_df
    total_demand = forecast_source["predicted_sales"].sum()
    average_daily_demand = total_demand / max(horizon_days, 1)

    confidence_interval_width = None
    if (
        "lower_bound" in forecast_source.columns
        and "upper_bound" in forecast_source.columns
    ):
        confidence_interval_width = float(
            (forecast_source["upper_bound"] - forecast_source["lower_bound"]).mean()
        )

    forecasting_metrics = {
        "horizon_days": horizon_days,
        "total_demand": round(total_demand, 0),
        "average_daily_demand": round(average_daily_demand, 2),
        "confidence_interval_width": round(confidence_interval_width, 2) if confidence_interval_width is not None else None,
        "future_forecast_df": forecast_source.copy(),
    }

    # ── Business metrics (inventory and replenishment decisions) ─────────────
    high_risk_count = (
        ((inventory_df["risk_level"] == "High Risk") | (inventory_df["risk_level"] == "Critical")).sum()
        if "risk_level" in inventory_df.columns
        else 0
    )
    medium_risk_count = (
        (inventory_df["risk_level"] == "Medium Risk").sum()
        if "risk_level" in inventory_df.columns
        else 0
    )
    low_risk_count = (
        (inventory_df["risk_level"] == "Low Risk").sum()
        if "risk_level" in inventory_df.columns
        else 0
    )

    top_combo = None
    if (
        "store_id" in forecast_source.columns
        and "dept_id" in forecast_source.columns
    ):
        top = (
            forecast_source.groupby(["store_id", "dept_id"])["predicted_sales"]
            .sum()
            .idxmax()
        )
        top_combo = f"Store {top[0]} / Dept {top[1]}"

    total_high_risk_demand = (
        inventory_df[(inventory_df["risk_level"] == "High Risk") | (inventory_df["risk_level"] == "Critical")]["forecast_demand"].sum()
        if "risk_level" in inventory_df.columns
        else 0
    )
    stockout_risk_pct = round(
        total_high_risk_demand / (total_demand + 1e-6) * 100, 1
    )

    avg_expected_stockout_probability = round(
        (
            inventory_df["expected_stockout_probability"].mean() * 100
            if "expected_stockout_probability" in inventory_df.columns
            else stockout_risk_pct
        ),
        1,
    )
    expected_overstock_total = round(
        inventory_df["expected_overstock_quantity"].sum()
        if "expected_overstock_quantity" in inventory_df.columns
        else 0.0,
        2,
    )
    average_days_of_supply = round(
        inventory_df["inventory_days_of_supply"].mean()
        if "inventory_days_of_supply" in inventory_df.columns
        else 0.0,
        2,
    )
    average_safety_stock_percentage = round(
        inventory_df["safety_stock_percentage"].mean()
        if "safety_stock_percentage" in inventory_df.columns
        else 0.0,
        2,
    )
    average_fill_rate = round(
        inventory_df["fill_rate_estimate"].mean()
        if "fill_rate_estimate" in inventory_df.columns
        else 0.0,
        2,
    )

    safety_stock_total = (
        inventory_df["safety_stock"].sum()
        if "safety_stock" in inventory_df.columns
        else 0.0
    )
    reorder_point_total = (
        inventory_df["reorder_point"].sum()
        if "reorder_point" in inventory_df.columns
        else 0.0
    )
    recommended_inventory_total = (
        inventory_df["recommended_inventory"].sum()
        if "recommended_inventory" in inventory_df.columns
        else 0.0
    )

    business_metrics = {
        "high_risk_combos": int(high_risk_count),
        "medium_risk_combos": int(medium_risk_count),
        "low_risk_combos": int(low_risk_count),
        "top_demand_combo": top_combo,
        "stockout_risk_pct": stockout_risk_pct,
        "expected_stockout_probability": avg_expected_stockout_probability,
        "expected_overstock_quantity": expected_overstock_total,
        "inventory_days_of_supply": average_days_of_supply,
        "safety_stock_percentage": average_safety_stock_percentage,
        "fill_rate_estimate": average_fill_rate,
        "total_skus": len(inventory_df),
        "safety_stock_total": round(safety_stock_total, 2),
        "reorder_point_total": round(reorder_point_total, 2),
        "recommended_inventory_total": round(recommended_inventory_total, 2),
        "inventory_df": inventory_df.copy(),
    }

    summary = {
        # evaluation (kept for UI backwards compat)
        "mape":                 mape,
        "rmse":                 rmse,
        "mae":                  mae,
        "baseline_improvement": baseline_improvement,
        # forecasting
        "total_demand":         round(total_demand, 0),
        "average_daily_demand": round(average_daily_demand, 2),
        "confidence_interval_width": round(confidence_interval_width, 2) if confidence_interval_width is not None else None,
        "horizon_days":         horizon_days,
        "model_name":           model_name,
        # business
        "high_risk_combos":     int(high_risk_count),
        "medium_risk_combos":   int(medium_risk_count),
        "low_risk_combos":      int(low_risk_count),
        "top_demand_combo":     top_combo,
        "stockout_risk_pct":    stockout_risk_pct,
        "expected_stockout_probability": avg_expected_stockout_probability,
        "expected_overstock_quantity":   expected_overstock_total,
        "inventory_days_of_supply":      average_days_of_supply,
        "safety_stock_percentage":       average_safety_stock_percentage,
        "fill_rate_estimate":            average_fill_rate,
        "total_skus":                    len(inventory_df),
        "safety_stock_total":            round(safety_stock_total, 2),
        "reorder_point_total":           round(reorder_point_total, 2),
        "recommended_inventory_total":   round(recommended_inventory_total, 2),
    }

    return {
        "evaluation_metrics": evaluation_metrics,
        "forecasting_metrics": forecasting_metrics,
        "business_metrics": business_metrics,
        "summary": summary,
    }


def export_excel(
    summary: dict,
    predictions_df: pd.DataFrame,
    inventory_df: pd.DataFrame,
    feature_importance_df: pd.DataFrame,
    results_df: pd.DataFrame,
    future_forecast_df: pd.DataFrame = None,
) -> bytes:
    """
    Build a formatted Excel workbook with 6 sheets and return as bytes.

    Sheets:
        1. Executive Summary
        2. Forecast Results
        3. Inventory Recommendations  (colour-coded by risk)
        4. Model Performance
        5. Feature Importance
        6. Future Forecast  (if provided)
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = openpyxl.Workbook()

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
    TITLE_FONT  = Font(bold=True, size=14, color="1F4E79")

    # ── Sheet 1: Executive Summary ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Executive Summary"

    ws["A1"] = "Retail Demand Forecasting — Executive Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")

    kpis = [
        ("Forecast Horizon",        f"{summary['horizon_days']} days"),
        ("Model Used",              summary["model_name"]),
        ("Total Forecast Demand",   f"{summary['total_demand']:,.0f} units"),
        ("Average Daily Demand",    f"{summary['average_daily_demand']:,.0f} units/day"),
        ("High Risk SKUs",          summary["high_risk_combos"]),
        ("Estimated Stockout Risk", f"{summary['stockout_risk_pct']}%"),
        ("Reorder Point",           f"{summary['reorder_point_total']:,.0f}"),
        ("Safety Stock",            f"{summary['safety_stock_total']:,.0f}"),
        ("Top Demand Segment",      summary["top_demand_combo"] or "N/A"),
    ]
    for i, (label, value) in enumerate(kpis, start=3):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value
        ws[f"A{i}"].font = Font(bold=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22

    # ── Sheet 2: Forecast Results ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Forecast Results")
    for r_idx, r in enumerate(
        dataframe_to_rows(predictions_df.round(2), index=False, header=True), start=1
    ):
        ws2.append(r)
        if r_idx == 1:
            for cell in ws2[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
    ws2.column_dimensions["A"].width = 14

    # ── Sheet 3: Inventory Recommendations ───────────────────────────────────
    ws3 = wb.create_sheet("Inventory Recommendations")
    HIGH_FILL = PatternFill("solid", fgColor="FFCCCC")
    MED_FILL  = PatternFill("solid", fgColor="FFF2CC")
    LOW_FILL  = PatternFill("solid", fgColor="CCFFCC")

    # Round only numeric cols — risk_level is categorical and will crash .round()
    inv_export = inventory_df.copy()
    num_cols = inv_export.select_dtypes(include="number").columns
    inv_export[num_cols] = inv_export[num_cols].round(2)
    # Defensively stringify any column openpyxl can't natively write (e.g. UUID)
    for col in inv_export.columns:
        if col not in num_cols and inv_export[col].dtype == "object":
            inv_export[col] = inv_export[col].apply(
                lambda v: str(v) if not isinstance(v, (str, type(None))) else v
            )

    for r_idx, r in enumerate(
        dataframe_to_rows(inv_export, index=False, header=True), start=1
    ):
        ws3.append(r)
        if r_idx == 1:
            for cell in ws3[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
        elif r_idx > 1 and "risk_level" in inventory_df.columns:
            risk_col = list(inventory_df.columns).index("risk_level") + 1
            risk_val = ws3.cell(row=r_idx, column=risk_col).value
            fill = {
                "High Risk":   HIGH_FILL,
                "Medium Risk": MED_FILL,
                "Low Risk":    LOW_FILL,
            }.get(str(risk_val))
            if fill:
                for cell in ws3[r_idx]:
                    cell.fill = fill

    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = 18

    # ── Sheet 4: Model Performance ────────────────────────────────────────────
    ws4 = wb.create_sheet("Model Performance")
    for r in dataframe_to_rows(results_df.round(4), index=False, header=True):
        ws4.append(r)
    for cell in ws4[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    # ── Sheet 5: Feature Importance ───────────────────────────────────────────
    ws5 = wb.create_sheet("Feature Importance")
    for r in dataframe_to_rows(
        feature_importance_df.head(20).round(4), index=False, header=True
    ):
        ws5.append(r)
    for cell in ws5[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    # ── Sheet 6: Future Forecast ──────────────────────────────────────────────
    if future_forecast_df is not None and len(future_forecast_df) > 0:
        ws6 = wb.create_sheet("Future Forecast")
        for r_idx, r in enumerate(
            dataframe_to_rows(future_forecast_df.round(2), index=False, header=True),
            start=1,
        ):
            ws6.append(r)
            if r_idx == 1:
                for cell in ws6[1]:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
        ws6.column_dimensions["A"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()